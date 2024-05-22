import sys
import openpyxl
import openpyxl.workbook
import re




####################################################################################################
##
## Modelo do catalogo - estrutura de entrada de dados
##
####################################################################################################
class Line:
    colunas = ['a', 'b', 'c', 'd']
    def __init__ (self, code, description, partnumber, ncm) :
        self.code : str = code
        self.description : str = description
        self.partnumber : str = partnumber
        self.ncm : str = ncm

    def read_line (i_linha : int, wb : openpyxl.workbook) :
        s_linha = ''
        planilha = wb['Planilha1']
        for coluna in Line.colunas :
            str_cel = coluna + str(i_linha)
            s_linha = str(s_linha) + str(planilha[str_cel].value).replace("\t", "    ") + '\t'

        split = s_linha.split("\t")

        line = Line(code = split[0],
                description = split[1],
                partnumber = split[2],
                ncm = split[3])
        
        return line




class Catalog:
    
    def __init__(self, filename : str) :
        self.workbook = openpyxl.load_workbook(filename)
        self.planilha = self.workbook['Planilha1']
        self.lines : list = []
        self.get_lines()

    def get_lines (self) :
        i_linha = 0
        while (True) :
            i_linha += 1

            line = Line.read_line(i_linha, self.workbook)

            if (line.code == '' or i_linha >500) :
                break

            self.lines.append(line)
        
        return





################################################################################################
##
## Estrutura de dados de componentes
##
################################################################################################

################################################################################################
######   aux parser   ##########################################################################

# substitui tags em templates de arquivos tipo S-Expression (kicad)
def fill_template (template_file_path, tag_and_value) :
    model_lines = open(template_file_path).readlines()
    model_reGex : str = r'<\S*>'
    result_lines : list = []
    for line in model_lines :
        for pair in tag_and_value :
            line = line.replace(pair[0],pair[1])
        result_lines.append(line)

    # check if all key was done
    for line in result_lines :
        result :list = []
        result = re.findall(model_reGex,line)
        if result.__len__() > 0 :
            print("Chave " + result[0] + " não encontrada para o modelo " + template_file_path)

    return result_lines


################################################################################################
######   resistores SMD   ######################################################################
class resistor_smd :
    def __init__ (self, line : Line) :
        self.code = line.code
        self.partnumber = line.partnumber
        self.description = line.description
        self.value = ''
        self.toleance = ''
        self.power = ''
        self.footprint = ''
        self.warining = ''
        self.error = ''
        self.kicad_model_lines = ['']


        # separa os valores da linha
        line_fields = line.description.split(" ")

        # Recupera os valores de resistência
        find :list = []
        for x in line_fields :
            if (re.findall(r'(\d+)([KRM])', x, re.IGNORECASE).__len__() > 0) :
                find.append(x)

        if (find.__len__() > 1) :
            self.error += ("Mais de um valor: " + line.description + " Encontrado: [" + "".join(x + ',' for x in find).removesuffix(",") + ']' )
            self.value = "".join(x + ',' for x in find).removesuffix(",")
        elif(find.__len__() == 0) :
            self.error+= ("Erro ao ler o valor: " + line.description + " Nada foi encontrado")
        else :
            self.value = str(find[0])

        if (self.value != '0R') :
            # recupera os valores de tolerância
            find :list = re.findall("\\b(\\d+[%])",line.description, re.IGNORECASE)
            find += re.findall("\\b(\\d+ppm)",line.description, re.IGNORECASE)
            if (find.__len__() > 1) :
                self.warining += ("Mais de uma tolerância encontrada: [" + "".join(x + ',' for x in find).removesuffix(",") + ']' )
                self.toleance = "".join(x + ',' for x in find).removesuffix(",")
            elif(find.__len__() == 0) :
                self.error+= ("Erro ao ler a tolerância: " + line.description + " Nada foi encontrado")
            else :
                self.toleance = find[0]

        # recupera os valores de potência
        find :list = re.findall("\\b(\\d/\\d+w) ",line.description, re.IGNORECASE) # captura o formato xx/xxW
        find += re.findall("\\s(\\d+w)",line.description, re.IGNORECASE)
        if (find.__len__() > 1) :
            self.warining += ("Mais de uma potência: " + line.description + " Encontrado: [" + "".join(x + ',' for x in find).removesuffix(",") + ']' )
            self.power = "".join(x + ',' for x in find).removesuffix(",")
        elif(find.__len__() == 0) :
            self.error+= ("Erro ao ler a potência: " + line.description + " Nada foi encontrado")
        else :
            self.power = find[0]

        # recupera os valores de footprint
        find :list = re.findall("0402|0603|0805|1206|2512|2010|0612",line.description, re.IGNORECASE)
        if (find.__len__() > 1) :
            self.warining += ("Mais de um footprint: " + line.description + " Encontrado: [" + "".join(x + ',' for x in find).removesuffix(",") + ']' )
            self.footprint = "".join(x + ',' for x in find).removesuffix(",")
        elif(find.__len__() == 0) :
            self.error+= ("Erro ao ler a footprint: " + line.description + " Nada foi encontrado")
        else :
            self.footprint = find[0]

        if (self.error == "" and self.warining == ""):
            self.template_file : str = 'models/symbol/resistor_smd/STANDARD.txt'
            self.tag_and_value = [  
                                ['<NAME>', 'RESISTOR SMD ' + str(self.value) + " " + str(self.toleance) + " " + str(self.footprint)],
                                ['<VALUE>', str(self.value)],
                                ['<SIZE>' , str(self.footprint)],
                                ['<POWER>' , str(self.power)],
                                ['<CODE>' , str(self.code)],
                                ['<TOLERANCE>', str(self.toleance)],
                                ['<DESCRIPTION>', str(self.description)]
                            ]
            self.kicad_model_lines = fill_template(self.template_file, self.tag_and_value)
            self.kicad_model_lines = fill_template(r'models/symbol/resistor_smd/lib_resistors.kicad_sym',['<INSERT>', str(self.kicad_model_lines)])


        

################################################################################################
#####   capacitores SMD   ######################################################################
class capacitor_smd :
    def __init__ (self, line : Line) :
        self.code = line.code
        self.partnumber = line.partnumber
        self.description = line.description
        self.value = ''
        self.toleance = ''
        self.tipe = ''
        self.footprint = ''
        self.voltage = ''
        self.warining = ''
        self.error = ''
        self.kicad_model_lines = ['']

        # Recupera os valores de capacitância
        find :list = re.findall("\\b(\\d+.*F)",line.description, re.IGNORECASE)
        if (find.__len__() > 1) :
            self.warining += ("Mais de um valor: " + line.description + " Encontrado: [" + "".join(x + ',' for x in find).removesuffix(",") + ']' )
            self.value = "".join(x + ',' for x in find).removesuffix(",")
        elif(find.__len__() == 0) :
            self.error+= ("Erro ao ler o valor: " + line.description + " Nada foi encontrado")
        else :
            self.value = find[0]

        # recupera os valores de tolerância
        find :list = re.findall("\\b(\\d+[%])",line.description, re.IGNORECASE)
        # find += re.findall("\\b(\\d+ppm)",line.description, re.IGNORECASE)
        if (find.__len__() > 1) :
            self.warining += ("Mais de uma tolerância: " + line.description + " Encontrado: [" + "".join(x + ',' for x in find).removesuffix(",") + ']' )
            self.toleance = "".join(x + ',' for x in find).removesuffix(",")
        elif(find.__len__() == 0) :
            self.warining += ("Atanção tolerância não encontrada: " + line.description)
        else :
            self.toleance = find[0]

        # recupera os valores de footprint
        find :list = re.findall("0402|0603|0805",line.description, re.IGNORECASE) #|1206|1210|2512|2010|0612|1812
        if (find.__len__() > 1) :
            self.warining += ("Mais de um footprint: " + line.description + " Encontrado: [" + "".join(x + ',' for x in find).removesuffix(",") + ']' )
            self.footprint = "".join(x + ',' for x in find).removesuffix(",")
        elif(find.__len__() == 0) :
            self.error+= ("Erro ao ler a footprint: " + line.description + " Nada foi encontrado")
        else :
            self.footprint = find[0]

        # recupera os valores de tipo
        find :list = re.findall("CER",line.description, re.IGNORECASE) # TAN|CER|ELE|SUPCAP <<<<<<<< alguns tipos são polarizados, tomar cuidado
        if (find.__len__() > 1) :
            self.error+= ("Erro mais de um tipo: " + line.description + " Encontrado: [" + "".join(x + ',' for x in find).removesuffix(",") + ']' )
        elif(find.__len__() == 0) :
            self.error+= ("Erro ao ler o tipo: " + line.description + " Nada foi encontrado")
        else :
            self.tipe = find[0]

        # Caso o tipo seja cerâmico, procura o tipo de isolante
        if (self.tipe == "CER") :
            find :list = re.findall("C0G|X7R|NPO|X5R",line.description, re.IGNORECASE)
            if (find.__len__() > 1) :
                self.warining += ("Atenção mais de um material: " + line.description + " Encontrado: [" + "".join(x + ',' for x in find).removesuffix(",") + ']' )
                self.tipe += " " + "".join(x + "/" for x in find).removesuffix("/")

            elif(find.__len__() == 0) :
                self.error += ("Erro ao ler o material: " + line.description + " Nada foi encontrado")
            else :
                self.tipe += " " + find[0]

        # obtém os valores de tensão
        find :list = re.findall(r'(\d+)([V])|(\d+)(kV)', line.description, re.IGNORECASE)
        if (find.__len__() > 1) :
            self.error+= ("Erro mais de uma tansão: " + line.description + " Encontrado: [" + "".join(x + ',' for x in find).removesuffix(",") + ']' )
        elif(find.__len__() == 0) :
            self.error+= ("Erro ao ler a tensão: " + line.description + " Nada foi encontrado")
        else :
            self.voltage = find[0]

        # finaliza preenchendo o template de arquivo
        if (self.error == "" and self.warining == ""):
            self.template_file : str = 'models/symbol/capacitor_smd/STANDARD.txt'
            self.tag_and_value = [  
                                ['<NAME>', 'CAPACITOR SMD ' + str(self.tipe) + " " + str(self.value) + " " + str(self.toleance) + " " + str(self.footprint)],
                                ['<VALUE>', str(self.value)],
                                ['<SIZE>' , str(self.footprint)],
                                ['<VOLTAGE>' , str(self.voltage)],
                                ['<CODE>' , str(self.code)],
                                ['<TOLERANCE>', str(self.toleance)],
                                ['<DESCRIPTION>', str(self.description)]
                            ]
            self.kicad_model_lines = fill_template(self.template_file, self.tag_and_value)

################################################################################################
####   lista de componentes   ##################################################################
class ComponentList :
    def __init__ (self) :
        self.workbook = openpyxl.Workbook()
        self.planilha = self.workbook[self.workbook.sheetnames[0]]

    def parse_lines (self, lines :Line = [] ):
        for line in lines:
            if re.search("001.101" ,line.code) :
                self.add_cap_smd(capacitor_smd(line))
            if re.search("001.102" ,line.code) :
                self.add_res_smd(resistor_smd(line))


    def addLine (self, line:Line) :
        str_cel = 'A' + str(self.line)
        self.planilha[str_cel].value = str(line.code)
        self.planilha.append([str(line.code), str(line.description), str(line.partnumber), str(line.ncm)])


    def add_res_smd (self, resistor:resistor_smd) :
        check_sheet = False
        for name in self.workbook.sheetnames :
            if (name == 'resistor_smd') :
                check_sheet = True
        if check_sheet == False :
            self.workbook.create_sheet('resistor_smd')
            lib = open('./out/resistors_smd.txt','w').close()

        self.planilha = self.workbook['resistor_smd']
        self.planilha.append([str(resistor.code), str(resistor.description), str(resistor.value), str(resistor.toleance), str(resistor.power), str(resistor.footprint), str(resistor.partnumber), str(resistor.warining), str(resistor.error),])
        lib = open('./out/resistors_smd.txt','a').writelines(resistor.kicad_model_lines)


    def add_cap_smd (self, capacitor:capacitor_smd) :
        check_sheet = False
        for name in self.workbook.sheetnames :
            if (name == 'capacitor_smd') :
                check_sheet = True
        if check_sheet == False :
            self.workbook.create_sheet('capacitor_smd')
            lib = open('./out/capacitors_smd.txt','w').close()
        
        self.planilha = self.workbook['capacitor_smd']
        self.planilha.append([str(capacitor.code), str(capacitor.description), str(capacitor.value), str(capacitor.toleance), str(capacitor.tipe), str(capacitor.footprint), str(capacitor.partnumber), str(capacitor.warining), str(capacitor.error)])
        lib = open('./out/capacitors_smd.txt','a').writelines(capacitor.kicad_model_lines)

    def save(self, name:str) :
        lib = open('./out/capacitors_smd.txt','r')
        self.kicad_model_lines = fill_template(r'models/symbol/capacitor_smd/lib_capacitors.kicad_sym',[['<INSERT>', "".join(lib.readlines())]])
        lib.close()
        lib = open('./out/capacitors_smd.kicad_sym','w')
        lib.writelines(self.kicad_model_lines)

        # lib = open('./out/resistor_smd.txt','r')
        # self.kicad_model_lines = fill_template(r'models/symbol/capacitor_smd/lib_resistors.kicad_sym',[['<INSERT>', "".join(lib.readlines())]])
        # lib.close()
        # lib = open('./out/resistors_smd.kicad_sym','w')
        # lib.writelines(self.kicad_model_lines)

        self.workbook.save(name)



################################################################################################
##
## aplicação
##
################################################################################################
def main () :
    catalog : Catalog = Catalog('./input/catalogo.xlsx')
    components : ComponentList = ComponentList()

    components.parse_lines(catalog.lines)
    components.save('./out/catalogo_smd.xlsx')

if __name__ == "__main__":
    main()